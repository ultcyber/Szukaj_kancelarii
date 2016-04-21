import requests
import re
from bs4 import BeautifulSoup
import openpyxl
import sys

'''
# For intro music
from selenium import webdriver
'''

# Load the page - verify if the page has information
def load_page(number):
	r = requests.get("http://www.szukajradcy.pl/Szukaj/Detail/rw_{}".format(number))
	if re.search("error occurred", r.text) is not None:
		return False
	else:
		return r.text

# Parse the page, set default value for cells ("brak"), return a tuple of 3 values for 3 columns

def parse_page(page):
	soup = BeautifulSoup(page,'html.parser')

	tags = soup.find_all("strong")
	kancelaria,strona,mail = "brak","brak","brak"

	for i in range(len(tags)):
		if re.search("Kancelaria",tags[i].string) is not None:
			kancelaria = tags[i].next_sibling.string
		elif re.search("WWW",tags[i].string) is not None:
			strona = tags[i].next_sibling.string
		elif re.search("mail",tags[i].string) is not None:
			mail = tags[i].next_sibling.string[::-1]
	
	return (kancelaria, strona, mail)


# create an xls file, make headers, insert data for every item in input argument
def create_xls(data_list):
	wb = openpyxl.Workbook()
	ws = wb.active
	row = 2
	ws['A1'], ws['B1'], ws['C1'] = "Kancelaria", "Strona", "Mail"
	for office in data_list:
		(ws['A'+str(row)], ws['B'+str(row)], ws['C'+str(row)]) = office
		row += 1
	wb.save("kancelarie.xls")


def main():
	'''
	# Intro music

	driver = webdriver.Firefox()
	driver.get("https://www.youtube.com/watch?v=tDq3fNew1rU")
	'''

	# A list of entries	
	list_of_offices = []

	# Start from 1, end range taken from the command line
	end_range = sys.argv[1]
	for i in range(1,int(end_range)):
		try:
			print("Analizuję wpis nr {}".format(i))
			# Continue only if page has info
			if load_page(i) != False:
				list_of_offices.append(parse_page(load_page(i)))
		except:
			print("Błąd dla wpisu nr {}".format(i))
			continue
		
	# Create an xls from the list
	create_xls(list_of_offices)

	return "Success!"

if __name__ == '__main__':
	main()
